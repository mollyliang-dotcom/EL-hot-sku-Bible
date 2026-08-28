#!/usr/bin/env python3
import argparse
import hashlib
import io
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image as PILImage


def parse_args():
    p = argparse.ArgumentParser(description="Extract text and floating images from Google Sheet XLSX.")
    p.add_argument("--xlsx", default="category_source.xlsx")
    p.add_argument("--sheet", default="category")
    p.add_argument("--out-dir", default=".category_sync_tmp")
    return p.parse_args()


def image_extension(raw: bytes) -> str:
    try:
        with PILImage.open(io.BytesIO(raw)) as im:
            fmt = (im.format or "PNG").lower()
            return "jpg" if fmt == "jpeg" else fmt
    except Exception:
        return "png"


def anchor_position(img):
    anchor = getattr(img, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return 1, 1
    return int(marker.row) + 1, int(marker.col) + 1


def choose_text_block(blocks, image_row, image_col):
    if not blocks:
        return None

    # Priority 1: text on the same row.
    same_row = [b for b in blocks if b["row"] == image_row]
    if same_row:
        return min(same_row, key=lambda b: abs(b["col"] - image_col))

    # Priority 2: nearby text above the image. This fits article-style sheets
    # where a paragraph/title is followed by a floating image.
    above = [b for b in blocks if b["row"] <= image_row]
    if above:
        return min(
            above,
            key=lambda b: (image_row - b["row"]) * 4 + abs(b["col"] - image_col),
        )

    # Fallback: nearest text block in the sheet.
    return min(
        blocks,
        key=lambda b: abs(b["row"] - image_row) * 4 + abs(b["col"] - image_col),
    )


def main():
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    out_dir = Path(args.out_dir)
    assets_dir = out_dir / "category_assets"
    json_path = out_dir / "category_content.json"

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    # Always build into a clean staging directory.
    if out_dir.exists():
        shutil.rmtree(out_dir)
    assets_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=False)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")

    ws = wb[args.sheet]

    blocks = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip():
                blocks.append({
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "text": value.strip(),
                    "images": [],
                })

    blocks.sort(key=lambda b: (b["row"], b["col"]))

    unmatched_images = []
    extracted_count = 0

    for index, img in enumerate(getattr(ws, "_images", []), start=1):
        row, col = anchor_position(img)

        try:
            raw = img._data()
        except Exception as exc:
            unmatched_images.append({
                "row": row,
                "col": col,
                "error": f"image read failed: {exc}",
            })
            continue

        if not raw:
            unmatched_images.append({
                "row": row,
                "col": col,
                "error": "empty image bytes",
            })
            continue

        ext = image_extension(raw)
        digest = hashlib.sha256(raw).hexdigest()[:12]
        filename = f"img_r{row}_c{col}_{digest}.{ext}"
        image_path = assets_dir / filename
        image_path.write_bytes(raw)
        extracted_count += 1

        rel_url = f"category_assets/{filename}"
        block = choose_text_block(blocks, row, col)

        if block is None:
            unmatched_images.append({
                "row": row,
                "col": col,
                "src": rel_url,
            })
        else:
            block["images"].append(rel_url)

    # Deduplicate images in each block while preserving order.
    for block in blocks:
        block["images"] = list(dict.fromkeys(block["images"]))

    payload = {
        "source": {
            "spreadsheet_id": "1h01w3EILyJbl0l7tSNXuvpDZ3kfuUfylYYXhgl7rVH4",
            "sheet": args.sheet,
        },
        "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "stats": {
            "text_blocks": len(blocks),
            "images_extracted": extracted_count,
            "images_unmatched": len(unmatched_images),
        },
        "blocks": blocks,
        "unmatched_images": unmatched_images,
    }

    # Validation: do not produce an "empty success" when the sheet unexpectedly fails.
    if not blocks and extracted_count == 0:
        raise RuntimeError("No text or floating images were extracted; refusing to publish empty content.")

    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Final self-check.
    check = json.loads(json_path.read_text(encoding="utf-8"))
    if "blocks" not in check or not isinstance(check["blocks"], list):
        raise RuntimeError("Generated JSON validation failed.")

    print(
        f"OK: {len(blocks)} text blocks, "
        f"{extracted_count} images, "
        f"{len(unmatched_images)} unmatched images"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

